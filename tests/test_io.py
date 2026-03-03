"""Tests for Excel I/O operations."""

from __future__ import annotations

from unittest.mock import patch

import openpyxl
import pandas as pd
import pytest

from case_parser.io import ExcelHandler, read_excel


class TestReadExcel:
    def test_reads_valid_xlsx(self, tmp_path):
        path = tmp_path / "test.xlsx"
        df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        df.to_excel(path, index=False)

        result = read_excel(path)

        assert list(result.columns) == ["A", "B"]
        assert len(result) == 2

    def test_raises_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            read_excel(tmp_path / "nonexistent.xlsx")

    def test_raises_value_error_for_bad_extension(self, tmp_path):
        path = tmp_path / "test.csv"
        path.write_text("a,b\n1,2")

        with pytest.raises(ValueError, match="Unsupported file format"):
            read_excel(path)

    def test_raises_type_error_for_multiple_sheets(self, tmp_path):
        path = tmp_path / "multi.xlsx"
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            pd.DataFrame({"A": [1]}).to_excel(writer, sheet_name="Sheet1", index=False)
            pd.DataFrame({"B": [2]}).to_excel(writer, sheet_name="Sheet2", index=False)

        with pytest.raises(TypeError, match="Multiple sheets"):
            read_excel(path, sheet_name=None)

    def test_reads_by_sheet_name(self, tmp_path):
        path = tmp_path / "named.xlsx"
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            pd.DataFrame({"X": [5, 6]}).to_excel(
                writer, sheet_name="MySheet", index=False
            )

        result = read_excel(path, sheet_name="MySheet")

        assert list(result.columns) == ["X"]
        assert len(result) == 2


class TestExcelHandlerWriteExcel:
    def test_writes_file_to_disk(self, tmp_path):
        handler = ExcelHandler()
        path = tmp_path / "output.xlsx"
        df = pd.DataFrame({"Col1": ["a", "b"], "Col2": [1, 2]})

        handler.write_excel(df, path)

        assert path.exists()
        result = pd.read_excel(path)
        assert len(result) == 2

    def test_creates_parent_directory(self, tmp_path):
        handler = ExcelHandler()
        path = tmp_path / "subdir" / "output.xlsx"
        df = pd.DataFrame({"A": [1]})

        handler.write_excel(df, path)

        assert path.exists()

    def test_sheet_name_respected(self, tmp_path):
        handler = ExcelHandler()
        path = tmp_path / "output.xlsx"
        df = pd.DataFrame({"A": [1]})

        handler.write_excel(df, path, sheet_name="MySheet")

        wb = openpyxl.load_workbook(path)
        assert "MySheet" in wb.sheetnames

    def test_raises_permission_error(self, tmp_path):
        handler = ExcelHandler()
        path = tmp_path / "output.xlsx"
        df = pd.DataFrame({"A": [1]})

        with (
            patch(
                "case_parser.io.pd.ExcelWriter", side_effect=PermissionError("locked")
            ),
            pytest.raises(PermissionError),
        ):
            handler.write_excel(df, path)

    def test_raises_on_generic_exception(self, tmp_path):
        handler = ExcelHandler()
        path = tmp_path / "output.xlsx"
        df = pd.DataFrame({"A": [1]})

        with (
            patch("case_parser.io.pd.ExcelWriter", side_effect=OSError("disk full")),
            pytest.raises(OSError, match="disk full"),
        ):
            handler.write_excel(df, path)


class TestAutoSizeColumns:
    def test_auto_width_capped_at_max(self, tmp_path):
        handler = ExcelHandler(max_width=20)
        path = tmp_path / "output.xlsx"
        df = pd.DataFrame({"A": ["x" * 100]})

        handler.write_excel(df, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["CaseLog"]
        assert ws.column_dimensions["A"].width <= 20

    def test_fixed_width_overrides_auto(self, tmp_path):
        handler = ExcelHandler()
        path = tmp_path / "output.xlsx"
        df = pd.DataFrame({"Col": ["a"]})

        handler.write_excel(df, path, fixed_widths={"Col": 42})

        wb = openpyxl.load_workbook(path)
        ws = wb["CaseLog"]
        assert ws.column_dimensions["A"].width == 42

    def test_auto_width_minimum_is_header_length(self, tmp_path):
        handler = ExcelHandler()
        path = tmp_path / "output.xlsx"
        # Short data (1 char) but long header → width driven by header
        df = pd.DataFrame({"LongHeaderName": ["x"]})

        handler.write_excel(df, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["CaseLog"]
        # Width should be at least len("LongHeaderName") = 14
        assert ws.column_dimensions["A"].width >= len("LongHeaderName")
